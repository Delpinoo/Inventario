from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .forms import ProductoForm, TelefonoForm
from .models import Producto, Sucursal, Telefono, Tipo
from django.conf import settings
from datetime import datetime
from openpyxl.cell.cell import Cell
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from openpyxl import Workbook
from django.contrib.auth import authenticate, login as auth_login, get_user_model , logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
import openpyxl
import os
import json



def escribir_valor(ws, fila, columna, valor):
    cell = ws.cell(row=fila, column=columna)
    if isinstance(cell, Cell) and not isinstance(cell, openpyxl.cell.cell.MergedCell):
        cell.value = valor
    else:
        #en esta parte se encuntra la celda principal cuando estan fucionadas
        for rango in ws.merged_cells.ranges:
            if (cell.coordinate in rango):
                superior_izquierda = rango.min_row, rango.min_col
                ws.cell(row=superior_izquierda[0], column=superior_izquierda[1]).value = valor
                break 
            

@login_required(login_url='login')
def home(request):
    productos = Producto.objects.none()
    form_agregar = ProductoForm()
    form_modificar = None
    form = None
    producto = None
    formulario_activo = None
    error = None
    sucursales = Sucursal.objects.all()

    sucursal_id = request.GET.get('sucursal')
    tipo_id = request.GET.get('tipo')

    productos = Producto.objects.all()

    if sucursal_id:
        productos = productos.filter(sucursal_id=sucursal_id)
    
    if tipo_id:
        productos = productos.filter(tipo_id=tipo_id)

    sucursales = Sucursal.objects.all()
    tipos = Tipo.objects.all() 



    # Agregar producto
    if request.method == 'POST' and 'agregar_producto' in request.POST:
        formulario_activo = 'agregar'
        form_agregar = ProductoForm(request.POST)
        if form_agregar.is_valid():
            producto = form_agregar.save(commit=False)
            
            
            sucursal_id = request.GET.get('sucursal', None)  
            if sucursal_id:
                producto.sucursal_id = sucursal_id  

            producto.save()
            return redirect('home')
    else:
        form_agregar = ProductoForm()

    if request.GET.get('exportar', False):  
        sucursal_id = request.GET.get('sucursal')
    
        if not sucursal_id:
            return HttpResponse("Error: No se ha proporcionado el ID de la sucursal", status=400)

        try:
            sucursal = Sucursal.objects.get(id=sucursal_id)
        except Sucursal.DoesNotExist:
            return HttpResponse(f"Error: No se encontró la sucursal con ID {sucursal_id}", status=404)
        
        plantilla_path = os.path.join(settings.BASE_DIR, 'app', 'static', 'templates', 'plantilla_inventario.xlsx')
        
        if not os.path.exists(plantilla_path):
            return HttpResponse(f"Error: El archivo plantilla no se encuentra en la ruta {plantilla_path}", status=404)

        wb = openpyxl.load_workbook(plantilla_path)
        ws = wb.active

        fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        escribir_valor(ws, 11, 12, fecha_actual)

        ws['L11'] = fecha_actual

        sucursal = Sucursal.objects.get(id=sucursal_id) 
    
        nombre_sucursal = sucursal.nombre_sucursal
        direccion = sucursal.direccion
        telefono = sucursal.telefono


        escribir_valor(ws, 16, 5, nombre_sucursal)
        escribir_valor(ws, 16, 8, direccion)
        escribir_valor(ws, 16, 11, telefono)
        
        productos = Producto.objects.filter(sucursal=sucursal)

        total = sum(producto.precio for producto in productos)      
        for row_num, producto in enumerate(productos, start=24):
            escribir_valor(ws, row_num, 4, producto.nombre)  
            escribir_valor(ws, row_num, 5, f"${producto.precio}")  

        escribir_valor(ws, 24, 6, f"${total}")
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=inventario_actualizado.xlsx'
        wb.save(response)
        print("Sucursal ID:", sucursal_id)
        print("Ruta de plantilla:", plantilla_path)
        return response
        
        
    return render(request, 'home/index.html', {
        'productos': productos,
        'form_agregar': form_agregar,
        'form': form,
        'producto': producto,
        'formulario_activo': formulario_activo,
        'error': error,
        'sucursales': sucursales,
        'tipos': tipos,
    })



def eliminar_productos(request):
    if request.method == 'POST':
        try:
            
            data = json.loads(request.body)
            ids_producto = data.get('ids', [])
            
            
            for producto_id in ids_producto:
                producto = Producto.objects.get(id=producto_id)
                producto.delete()
            
            
            return JsonResponse({'success': True})
        except Exception as e:
            
            return JsonResponse({'success': False, 'error': str(e)})


def modificar_producto(request):
    if request.method == 'POST':
        producto_id = request.POST.get('id')
        nombre = request.POST.get('nombre')
        numero_serie = request.POST.get('numero_serie')
        precio = request.POST.get('precio')
        cantidad = request.POST.get('cantidad')
        tipo_nombre = request.POST.get('tipo')  
        sucursal = request.POST.get('sucursal')

        producto = get_object_or_404(Producto, id=producto_id)
        tipo = get_object_or_404(Tipo, nombre=tipo_nombre)  
        sucursal = get_object_or_404(Sucursal, nombre_sucursal=sucursal)

        # Actualizar los campos del producto
        producto.nombre = nombre
        producto.numero_serie = numero_serie
        producto.precio = precio
        producto.cantidad = cantidad
        producto.tipo = tipo  
        producto.sucursal = sucursal
        producto.save()

        return redirect('home')  
    
    
    tipos = Tipo.objects.all()  
    sucursales = Sucursal.objects.all()  
    
    return render(request, 'phones/telefonos.html', {
        "tipos": tipos, 
        "sucursales": sucursales})

#--------------------------------------------------------------------------------------------------------------------
def escribir_valor_en_celda(ws, fila, columna, valor):
    cell = ws.cell(row=fila, column=columna)
    if isinstance(cell, Cell) and not isinstance(cell, openpyxl.cell.cell.MergedCell):
        cell.value = valor
    else:
        
        for rango in ws.merged_cells.ranges:
            if (cell.coordinate in rango):
                superior_izquierda = rango.min_row, rango.min_col
                ws.cell(row=superior_izquierda[0], column=superior_izquierda[1]).value = valor
                break  

@login_required(login_url='login')
def phones(request):
    
    if request.method == "POST":
        nombre_dueño = request.POST.get("nombre_dueño")
        modelo_telefono = request.POST.get("modelo_telefono")
        fono = request.POST.get("fono")
        sucursal_id = request.POST.get("sucursal")

        
        if nombre_dueño and modelo_telefono and fono and sucursal_id:
            sucursal = Sucursal.objects.get(id=sucursal_id)
            Telefono.objects.create(
                nombre_dueño=nombre_dueño,
                modelo_telefono=modelo_telefono,
                fono=fono,
                sucursal=sucursal
            )
        return redirect("phones")  

    
    telefonos = Telefono.objects.all()

    
    sucursal_id = request.GET.get('sucursal')
    if sucursal_id:
        telefonos = telefonos.filter(sucursal_id=sucursal_id)

    
    telefonos_por_sucursal = {}
    for telefono in telefonos:
        sucursal = telefono.sucursal
        if sucursal not in telefonos_por_sucursal:
            telefonos_por_sucursal[sucursal] = []
        telefonos_por_sucursal[sucursal].append(telefono)

    
    sucursales = Sucursal.objects.all()


    
    if request.GET.get('exportar') == 'true':
        
        plantilla_path = os.path.join(settings.BASE_DIR, 'app', 'static', 'templates', 'Inventario_Telefonos.xlsx')

        
        if not os.path.exists(plantilla_path):
            return HttpResponse(f"Error: El archivo plantilla no se encuentra en la ruta {plantilla_path}", status=404)

        
        wb = openpyxl.load_workbook(plantilla_path)
        ws = wb.active

        
        fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        escribir_valor_en_celda(ws, 11, 12, fecha_actual)  
        ws['L11'] = fecha_actual

        
        sucursal = Sucursal.objects.get(id=sucursal_id)
        nombre_sucursal = sucursal.nombre_sucursal
        direccion = sucursal.direccion
        telefono = sucursal.telefono

        
        escribir_valor_en_celda(ws, 38, 5, nombre_sucursal)  
        escribir_valor_en_celda(ws, 38, 8, direccion)  
        escribir_valor_en_celda(ws, 38, 11, telefono)  
        escribir_valor_en_celda(ws, 14, 7, nombre_sucursal)
        
        telefonos = Telefono.objects.filter(sucursal=sucursal)

        
        for row_num, telefono in enumerate(telefonos, start=14):
            escribir_valor_en_celda(ws, row_num, 4, telefono.nombre_dueño)  
            escribir_valor_en_celda(ws, row_num, 5, telefono.modelo_telefono)  
            escribir_valor_en_celda(ws, row_num, 6, telefono.fono)   

        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=inventario_actualizado.xlsx'
        
        
        wb.save(response)

        print("Sucursal ID:", sucursal_id)
        print("Ruta de plantilla:", plantilla_path)

        return response

    return render(request, 'phones/telefonos.html', {
        'telefonos_por_sucursal': telefonos_por_sucursal,
        'sucursales': sucursales,
        'request': request
    })




def modificar_telefono(request, telefono_id):
    telefono = get_object_or_404(Telefono, id=telefono_id)

    if request.method == 'POST':
        nombre_dueño = request.POST.get('nombre_dueño')
        modelo_telefono = request.POST.get('modelo_telefono')
        fono = request.POST.get('fono')
        sucursal_id = request.POST.get('sucursal')

        if nombre_dueño and modelo_telefono and fono and sucursal_id:
            telefono.nombre_dueño = nombre_dueño
            telefono.modelo_telefono = modelo_telefono
            telefono.fono = fono
            telefono.sucursal_id = sucursal_id
            telefono.save()
            return redirect('phones')  

    return JsonResponse({'error': 'Solicitud inválida'}, status=400)

@csrf_exempt
def eliminar_telefonos(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            ids = data.get('ids', [])
            if ids:
                
                print(f'IDs recibidos para eliminación: {ids}')
                
                
                telefonos_a_eliminar = Telefono.objects.filter(id__in=ids)
                
                
                if telefonos_a_eliminar.exists():
                    telefonos_a_eliminar.delete()
                    return JsonResponse({'success': True})
                else:
                    return JsonResponse({'success': False, 'message': 'No se encontraron teléfonos con esos IDs'})
            else:
                return JsonResponse({'success': False, 'message': 'No se proporcionaron IDs'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

    return JsonResponse({'success': False, 'message': 'Método de solicitud no permitido'}, status=400)


User = get_user_model()

def login_function(request):  
    if request.method == "POST":
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "").strip()

        if not email or not password:
            messages.error(request, "Debes ingresar un correo y una contraseña.")
            return redirect("login")

        
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            messages.error(request, "No se encontró un usuario con ese correo.")
            return redirect("login")

        
        user = authenticate(request, username=email, password=password)

        if user is not None:
            auth_login(request, user)
            messages.success(request, "Inicio de sesión exitoso.")
            return redirect("home")
        else:
            messages.error(request, "Correo o contraseña incorrectos.")
            return redirect("login")

    return render(request, "login/login.html") 

def logout_function(request):
    logout(request)
    return redirect("login")  